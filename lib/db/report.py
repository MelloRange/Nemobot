import os
from openpyxl import Workbook
import discord

from apscheduler.triggers.cron import CronTrigger
import datetime

from . import db

RP_PATH = "./data/db/weekly_report.xlsx"
wb = Workbook()
ws = wb.active
wb.save(RP_PATH)


def create_workbook():
	column = []
	zeros = []
	for name in db.get_column('SELECT name FROM Mods'):
		column.append(name)
		zeros.append(0)
	add_column(column)
	for x in range(24):
		add_column(zeros)

	wb.save(RP_PATH)

def add_column(column):
    new_column = ws.max_column + 1
    
    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

def add_to_column(hr, column):
	for rowy, value in enumerate(column, start=1):
		preval = ws.cell(row=rowy, column=hr).value
		if preval != None:
			ws.cell(row=rowy, column=hr, value=preval + value)

def update_hour():
	now = datetime.datetime.now()
	hr = now.hour
	column = []
	for mod in db.get_all('SELECT * FROM Mods'):
		if mod[2]:
			column.append(1)
		else:
			column.append(0)
	add_to_column(hr, column)
	wb.save(RP_PATH)
	return

def save_weekly_report():
	if os.path.exists(RP_PATH):
		if os.path.exists("./data/db/send_report.xlsx"):
			os.remove("./data/db/send_report.xlsx")
		os.rename(RP_PATH, "./data/db/send_report.xlsx")
	create_workbook()
	return

def send_report(sched):
	sched.add_job(save_weekly_report, CronTrigger(day_of_week=0))
	sched.add_job(update_hour, CronTrigger(minute=0, jitter=5))
